import openpyxl, sys

from openpyxl.utils import get_column_letter
from openpyxl.styles import Font


number= int(sys.argv[1])

wb=openpyxl.Workbook()
sheet=wb.active
boldfont=Font(bold=True)

 
for i in range(1,number+1):
    sheet['A'+str(i+1)]=i
    sheet['A'+str(i+1)].font=boldfont

for i in range (1,number+1):
    sheet[get_column_letter(i+1)+ str(1)]=i
    sheet[get_column_letter(i+1)+ str(1)].font=boldfont
for i in range(1,number+1):
    for j in range(1,number+1):
        sheet[get_column_letter(i+1)+str(j+1)]=i*j

wb.save('multiplication_table.xlsx')
