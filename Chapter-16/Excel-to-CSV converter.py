import openpyxl, csv, os
for files in os.listdir('.'):
    if files.endswith('.xlsx'):
        wb=openpyxl.load_workbook(files)
        
        for sheets in wb.get_sheet_names():
            sheet=wb.get_sheet_by_name(sheets)
            nameget=files[:-5]
            csvfile=open(nameget+'_'+sheets+'.csv','w',newline='')
            csvwrite=csv.writer(csvfile)

            for rows in range (1,sheet.max_row + 1):
                rowdata=[]
                for cols in range(1, sheet.max_column+1):
                    rowdata.append(sheet.cell(row=rows,column=cols).value)
                for row in rowdata:
                    csvwrite.writerow(row)
            csvfile.close()

        
            
